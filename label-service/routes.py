"""API routes for label management."""
import csv
import logging
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Optional, List, Dict, Any
from uuid import UUID
import httpx
from fastapi import APIRouter, Depends, HTTPException, Query, status, UploadFile, File, BackgroundTasks
from fastapi.responses import JSONResponse
import json

from openpyxl import load_workbook

from database import get_db
from crud import LabelCRUD, FeedbackSentimentCRUD, FeedbackIntentCRUD
from gemini_service import get_gemini_service
from schemas import (
    LabelCreate,
    LabelUpdate,
    LabelResponse,
    LabelTreeResponse,
    LabelListResponse,
    HealthResponse,
    BulkLabelCreate,
    BulkLabelResponse,
    FeedbackSentimentCreate,
    FeedbackSentimentUpdate,
    FeedbackSentimentResponse,
    FeedbackSentimentListResponse,
    FeedbackSource,
    SentimentLabel,
    IntentTriplet,
    FeedbackIntentResponse,
    LabelSyncItem,
    LabelSyncRequest,
    LabelSyncResponse,
    FeedbackReprocessStatus,
    FeedbackImportResponse,
)
from config import get_settings
from training_manager import get_training_manager
from training_log import is_training_in_progress

logger = logging.getLogger(__name__)
settings = get_settings()

router = APIRouter()


@router.get("/health", response_model=HealthResponse, summary="Health check endpoint")
def health_check():
    """Health check endpoint."""
    return HealthResponse(
        status="healthy",
        timestamp=datetime.utcnow(),
        database="ok",
        version=settings.app_version
    )


@router.get(
    "/training/status",
    summary="Ki·ªÉm tra tr·∫°ng th√°i trigger hu·∫•n luy·ªán"
)
async def training_status():
    """Ki·ªÉm tra tr·∫°ng th√°i training c·ªßa sentiment v√† embedding services."""
    try:
        manager = get_training_manager()
    except RuntimeError:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Training manager ch∆∞a ƒë∆∞·ª£c kh·ªüi t·∫°o"
        )
    
    # Ki·ªÉm tra tr·∫°ng th√°i training
    sentiment_training = is_training_in_progress("sentiment")
    embedding_training = is_training_in_progress("embedding")
    
    status_info = {
        "sentiment": {
            "is_training": sentiment_training,
            "message": "üîÑ ƒêang training sentiment model" if sentiment_training else "‚úÖ Sentiment model s·∫µn s√†ng"
        },
        "embedding": {
            "is_training": embedding_training,
            "message": "üîÑ ƒêang training embedding model" if embedding_training else "‚úÖ Embedding model s·∫µn s√†ng"
        },
        "warning": None
    }
    
    # Th√™m c·∫£nh b√°o n·∫øu c·∫£ hai ƒëang training
    if sentiment_training and embedding_training:
        status_info["warning"] = "‚ö†Ô∏è  C·∫£ sentiment v√† embedding ƒëang training - c√°c service c√≥ th·ªÉ b·ªã ch·∫≠m"
    elif sentiment_training:
        status_info["warning"] = "‚ö†Ô∏è  ƒêang training sentiment - embedding service c√≥ th·ªÉ b·ªã ch·∫≠m"
    elif embedding_training:
        status_info["warning"] = "‚ö†Ô∏è  ƒêang training embedding - sentiment service c√≥ th·ªÉ b·ªã ch·∫≠m"
    
    return status_info


@router.post(
    "/labels",
    response_model=LabelResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Create a new label"
)
async def create_label(label_data: LabelCreate, background_tasks: BackgroundTasks):
    """Create a new label in the hierarchy."""
    try:
        with get_db() as conn:
            # Check if label with same name and parent already exists
            exists = LabelCRUD.exists_by_name_and_parent(
                conn, label_data.name, label_data.parent_id
            )
            if exists:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Label '{label_data.name}' already exists under this parent"
                )
            
            label = LabelCRUD.create(conn, label_data)

            # Trigger training check in the background if enabled
            if settings.enable_training_trigger:
                background_tasks.add_task(trigger_training_on_new_label)

            return label
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except Exception as e:
        logger.error(f"Error creating label: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to create label"
        )

async def trigger_training_on_new_label():
    """Background task to trigger training check after a new label is created."""
    try:
        manager = await get_training_manager()
        await manager.check_and_trigger_all(triggered_by='new_label')
    except Exception as e:
        logger.error(f"Failed to trigger training on new label: {e}", exc_info=True)


@router.post(
    "/labels/bulk",
    response_model=BulkLabelResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Create multiple labels at once"
)
def create_labels_bulk(bulk_data: BulkLabelCreate):
    """Create multiple labels in a single request.
    
    This endpoint allows you to create multiple labels at once.
    Each label is validated and created independently.
    If some labels fail, others will still be created successfully.
    """
    try:
        with get_db() as conn:
            results = LabelCRUD.bulk_create(conn, bulk_data.labels)
            
            # Count successes and failures
            successful = sum(1 for r in results if r['success'])
            failed = sum(1 for r in results if not r['success'])
            
            return BulkLabelResponse(
                total=len(bulk_data.labels),
                successful=successful,
                failed=failed,
                results=results
            )
    except Exception as e:
        logger.error(f"Error in bulk label creation: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to process bulk label creation"
        )


@router.get(
    "/labels/export",
    summary="Export all labels as JSON backup"
)
def export_labels_backup():
    """Export t·∫•t c·∫£ labels ra file JSON ƒë·ªÉ backup."""
    try:
        with get_db() as conn:
            # L·∫•y t·∫•t c·∫£ labels kh√¥ng gi·ªõi h·∫°n
            labels = LabelCRUD.get_all(conn, skip=0, limit=100000)
            
            # Chuy·ªÉn ƒë·ªïi sang format LabelSyncItem (c√≥ id, name, level, parent_id, description, updated_at)
            export_data = {
                "labels": [
                    {
                        "id": label["id"],
                        "name": label["name"],
                        "level": label["level"],
                        "parent_id": label.get("parent_id"),
                        "description": label.get("description"),
                        "updated_at": label.get("updated_at").isoformat() if label.get("updated_at") else None
                    }
                    for label in labels
                ],
                "exported_at": datetime.utcnow().isoformat(),
                "total": len(labels)
            }
            
            return JSONResponse(
                content=export_data,
                headers={
                    "Content-Disposition": f"attachment; filename=labels_backup_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
                }
            )
    except Exception as e:
        logger.error(f"Error exporting labels: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to export labels"
        )


@router.post(
    "/labels/sync",
    response_model=LabelSyncResponse,
    summary="Sync labels from external system (JSON file or JSON body)"
)
async def sync_labels(
    sync_request: Optional[LabelSyncRequest] = None,
    file: Optional[UploadFile] = File(None)
):
    """ƒê·ªìng b·ªô label t·ª´ h·ªá th·ªëng ngo√†i v√† x·ª≠ l√Ω feedback b·ªã ·∫£nh h∆∞·ªüng.
    
    C√≥ th·ªÉ nh·∫≠n d·ªØ li·ªáu t·ª´:
    - File JSON upload (file parameter)
    - JSON body (sync_request parameter)
    """
    impacted_feedback_ids: List[UUID] = []
    labels_to_sync = []

    try:
        # N·∫øu c√≥ file upload, ƒë·ªçc t·ª´ file
        if file:
            if not file.filename or not file.filename.lower().endswith('.json'):
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Ch·ªâ h·ªó tr·ª£ file JSON"
                )
            content = await file.read()
            try:
                data = json.loads(content.decode('utf-8'))
                # H·ªó tr·ª£ c·∫£ format c√≥ "labels" key ho·∫∑c tr·ª±c ti·∫øp l√† array
                if isinstance(data, dict) and "labels" in data:
                    labels_data = data["labels"]
                elif isinstance(data, list):
                    labels_data = data
                else:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="File JSON kh√¥ng ƒë√∫ng ƒë·ªãnh d·∫°ng. C·∫ßn c√≥ key 'labels' ho·∫∑c l√† array"
                    )
                
                # Chuy·ªÉn ƒë·ªïi dicts th√†nh LabelSyncItem objects
                labels_to_sync = []
                for label_dict in labels_data:
                    try:
                        # Parse updated_at n·∫øu c√≥ (c√≥ th·ªÉ l√† string ISO format)
                        if "updated_at" in label_dict and label_dict["updated_at"]:
                            if isinstance(label_dict["updated_at"], str):
                                try:
                                    label_dict["updated_at"] = datetime.fromisoformat(label_dict["updated_at"].replace('Z', '+00:00'))
                                except ValueError:
                                    # Fallback n·∫øu format kh√¥ng ƒë√∫ng
                                    label_dict["updated_at"] = None
                        label_item = LabelSyncItem(**label_dict)
                        labels_to_sync.append(label_item)
                    except Exception as e:
                        raise HTTPException(
                            status_code=status.HTTP_400_BAD_REQUEST,
                            detail=f"Label kh√¥ng h·ª£p l·ªá: {label_dict.get('id', 'unknown')} - {str(e)}"
                        )
            except json.JSONDecodeError as e:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"File JSON kh√¥ng h·ª£p l·ªá: {str(e)}"
                )
        elif sync_request:
            # N·∫øu c√≥ sync_request t·ª´ body
            labels_to_sync = sync_request.labels
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="C·∫ßn cung c·∫•p file JSON ho·∫∑c JSON body"
            )
        
        if not labels_to_sync:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Kh√¥ng c√≥ label n√†o ƒë·ªÉ sync"
            )

        with get_db() as conn:
            sync_result = LabelCRUD.sync_labels(conn, labels_to_sync)
            changed_label_ids = sync_result.get("changed_label_ids", [])

            if changed_label_ids:
                impacted_feedback_ids = FeedbackSentimentCRUD.reset_feedback_for_labels(
                    conn,
                    changed_label_ids,
                )
                if impacted_feedback_ids:
                    FeedbackIntentCRUD.delete_by_feedback_ids(conn, impacted_feedback_ids)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(exc)
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Error syncing labels: %s", exc, exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to sync labels"
        )

    reprocess_results = await _reprocess_feedbacks_after_label_update(impacted_feedback_ids)

    return LabelSyncResponse(
        created=sync_result["created"],
        updated=sync_result["updated"],
        unchanged=sync_result["unchanged"],
        results=sync_result["results"],
        total=len(labels_to_sync),
        changed_label_ids=sync_result.get("changed_label_ids", []),
        impacted_feedbacks=len(impacted_feedback_ids),
        reprocessed_feedbacks=reprocess_results,
    )


@router.get(
    "/labels",
    response_model=LabelListResponse,
    summary="Get all labels with optional filters"
)
def get_labels(
    level: Optional[int] = Query(None, ge=1, le=3, description="Filter by level"),
    parent_id: Optional[int] = Query(None, ge=1, description="Filter by parent ID"),
    skip: int = Query(0, ge=0, description="Number of records to skip"),
    limit: int = Query(100, ge=1, le=1000, description="Maximum number of records"),
):
    """Get all labels with optional filters."""
    try:
        with get_db() as conn:
            labels = LabelCRUD.get_all(conn, level=level, parent_id=parent_id, skip=skip, limit=limit)
            total = LabelCRUD.count(conn, level=level, parent_id=parent_id)
            
            return LabelListResponse(
                labels=labels,
                total=total,
                level=level,
                parent_id=parent_id
            )
    except Exception as e:
        logger.error(f"Error getting labels: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to retrieve labels"
        )


@router.get(
    "/labels/tree",
    response_model=list[LabelTreeResponse],
    summary="Get full label hierarchy as tree"
)
def get_label_tree():
    """Get the full label hierarchy as a tree structure."""
    try:
        with get_db() as conn:
            tree = LabelCRUD.get_tree(conn)
            return tree
    except Exception as e:
        logger.error(f"Error getting label tree: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to retrieve label tree"
        )


@router.get(
    "/labels/{label_id}",
    response_model=LabelResponse,
    summary="Get label by ID"
)
def get_label(label_id: int):
    """Get a specific label by ID."""
    try:
        with get_db() as conn:
            label = LabelCRUD.get_by_id(conn, label_id)
            if not label:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Label with id {label_id} not found"
                )
            return label
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting label {label_id}: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to retrieve label"
        )


@router.get(
    "/labels/{label_id}/children",
    response_model=list[LabelResponse],
    summary="Get children of a label"
)
def get_label_children(label_id: int):
    """Get all children of a specific label."""
    try:
        with get_db() as conn:
            # First check if parent exists
            parent = LabelCRUD.get_by_id(conn, label_id)
            if not parent:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Label with id {label_id} not found"
                )
            
            children = LabelCRUD.get_children(conn, label_id)
            return children
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting children for label {label_id}: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to retrieve label children"
        )


@router.put(
    "/labels/{label_id}",
    response_model=LabelResponse,
    summary="Update a label"
)
def update_label(
    label_id: int,
    label_data: LabelUpdate,
):
    """Update a label's name or description."""
    try:
        with get_db() as conn:
            label = LabelCRUD.update(conn, label_id, label_data)
            if not label:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Label with id {label_id} not found"
                )
            return label
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating label {label_id}: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to update label"
        )


@router.delete(
    "/labels/{label_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Delete a label"
)
def delete_label(label_id: int):
    """Delete a label and all its children (cascade delete)."""
    try:
        with get_db() as conn:
            deleted = LabelCRUD.delete(conn, label_id)
            if not deleted:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Label with id {label_id} not found"
                )
            return None
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting label {label_id}: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to delete label"
        )


# --- Feedback Sentiment Routes ---

async def call_sentiment_service(text: str) -> dict:
    """Call the sentiment analysis service to classify text."""
    # Ki·ªÉm tra xem c√≥ ƒëang training embedding kh√¥ng
    if is_training_in_progress("embedding"):
        logger.warning("‚ö†Ô∏è  ƒêang training embedding model - sentiment service c√≥ th·ªÉ b·ªã ch·∫≠m ho·∫∑c kh√¥ng kh·∫£ d·ª•ng")
    
    # Ki·ªÉm tra xem c√≥ ƒëang training sentiment kh√¥ng
    if is_training_in_progress("sentiment"):
        logger.warning("üîÑ ƒêang training sentiment model - sentiment service c√≥ th·ªÉ kh√¥ng kh·∫£ d·ª•ng, y√™u c·∫ßu s·∫Ω ph·∫£i ch·ªù")
    
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.post(
                f"{settings.sentiment_service_url}/classify",
                json={"text": text}
            )
            response.raise_for_status()
            return response.json()
    except httpx.HTTPError as e:
        logger.error(f"Error calling sentiment service: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Sentiment analysis service is unavailable"
        )
    except Exception as e:
        logger.error(f"Unexpected error calling sentiment service: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to analyze sentiment"
        )


async def call_embedding_service(text: str) -> list:
    """Call the embedding service to get text embedding."""
    # Ki·ªÉm tra xem c√≥ ƒëang training sentiment kh√¥ng
    if is_training_in_progress("sentiment"):
        logger.warning("‚ö†Ô∏è  ƒêang training sentiment model - embedding service c√≥ th·ªÉ b·ªã ch·∫≠m ho·∫∑c kh√¥ng kh·∫£ d·ª•ng")
    
    # Ki·ªÉm tra xem c√≥ ƒëang training embedding kh√¥ng
    if is_training_in_progress("embedding"):
        logger.warning("üîÑ ƒêang training embedding model - embedding service c√≥ th·ªÉ kh√¥ng kh·∫£ d·ª•ng, y√™u c·∫ßu s·∫Ω ph·∫£i ch·ªù")
    
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.post(
                f"{settings.embedding_service_url}/encode",
                json={"text": text}
            )
            response.raise_for_status()
            result = response.json()
            return result.get("embedding", [])
    except httpx.HTTPError as e:
        logger.error(f"Error calling embedding service: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Embedding service is unavailable"
        )
    except Exception as e:
        logger.error(f"Unexpected error calling embedding service: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to get embedding"
        )


async def _classify_intent_for_feedback(
    feedback_text: str,
    feedback_id: Optional[UUID] = None
) -> Optional[Dict[str, Any]]:
    """
    Helper function to classify intent for a feedback.
    Returns dict with level1_id, level2_id, level3_id if successful, None otherwise.
    """
    try:
        # Step 1: Get embedding for feedback
        feedback_embedding = await call_embedding_service(feedback_text)
        
        if not feedback_embedding:
            logger.warning(f"Failed to get embedding for feedback {feedback_id or 'new'}")
            return None
        
        # Step 2: Get top 10 intent candidates
        with get_db() as conn:
            intent_candidates = FeedbackIntentCRUD.get_top_intents(
                conn,
                feedback_embedding,
                limit=10,
                top_level1=5,
                top_level2_total=15,
                top_level3_total=50
            )
            
            if not intent_candidates:
                logger.warning(f"No intent candidates found for feedback {feedback_id or 'new'}")
                return None
            
            # Step 3: Use Gemini to select best intent from top 10
            try:
                gemini_service = get_gemini_service()
                selected_intent = gemini_service.select_best_intent(
                    feedback_text,
                    intent_candidates
                )
                if selected_intent:
                    return {
                        'level1_id': selected_intent['level1']['id'],
                        'level2_id': selected_intent['level2']['id'],
                        'level3_id': selected_intent['level3']['id']
                    }
            except Exception as e:
                logger.warning(f"Gemini service error for feedback {feedback_id or 'new'}: {e}")
                return None
        
        return None
    except Exception as e:
        logger.error(f"Error classifying intent for feedback {feedback_id or 'new'}: {e}", exc_info=True)
        return None


async def _reprocess_feedbacks_after_label_update(
    feedback_ids: List[UUID],
) -> List[dict]:
    """Re-run intent prediction pipeline for impacted feedbacks."""
    results: List[dict] = []
    if not feedback_ids:
        return results

    try:
        gemini_service = get_gemini_service()
    except Exception as exc:  # pragma: no cover - initialization may fail in env
        logger.error("Gemini service unavailable for reprocess: %s", exc, exc_info=True)
        gemini_service = None

    for feedback_id in feedback_ids:
        result = {
            "feedback_id": feedback_id,
            "status": FeedbackReprocessStatus.SKIPPED.value,
            "message": None,
        }

        try:
            with get_db() as conn:
                feedback = FeedbackSentimentCRUD.get_by_id(conn, feedback_id)
        except Exception as exc:
            logger.error("Failed to load feedback %s: %s", feedback_id, exc, exc_info=True)
            result["status"] = FeedbackReprocessStatus.FAILED.value
            result["message"] = "Kh√¥ng ƒë·ªçc ƒë∆∞·ª£c d·ªØ li·ªáu feedback"
            results.append(result)
            continue

        if not feedback:
            result["status"] = FeedbackReprocessStatus.FAILED.value
            result["message"] = "Feedback kh√¥ng t·ªìn t·∫°i"
            results.append(result)
            continue

        try:
            embedding = await call_embedding_service(feedback["feedback_text"])
        except HTTPException as exc:
            logger.error(
                "Embedding service error (feedback %s): %s",
                feedback_id,
                exc.detail,
            )
            result["status"] = FeedbackReprocessStatus.FAILED.value
            result["message"] = f"L·ªói embedding: {exc.detail}"
            results.append(result)
            continue
        except Exception as exc:
            logger.error(
                "Unexpected embedding error (feedback %s): %s",
                feedback_id,
                exc,
                exc_info=True,
            )
            result["status"] = FeedbackReprocessStatus.FAILED.value
            result["message"] = "L·ªói b·∫•t ng·ªù khi g·ªçi embedding service"
            results.append(result)
            continue

        if not embedding:
            result["status"] = FeedbackReprocessStatus.SKIPPED.value
            result["message"] = "Embedding r·ªóng"
            results.append(result)
            continue

        try:
            with get_db() as conn:
                intent_candidates = FeedbackIntentCRUD.get_top_intents(
                    conn,
                    embedding,
                    limit=10,
                    top_level1=5,
                    top_level2_total=15,
                    top_level3_total=50,
                )
        except Exception as exc:
            logger.error(
                "Failed to compute intent candidates (feedback %s): %s",
                feedback_id,
                exc,
                exc_info=True,
            )
            result["status"] = FeedbackReprocessStatus.FAILED.value
            result["message"] = "Kh√¥ng t√≠nh ƒë∆∞·ª£c intent candidates"
            results.append(result)
            continue

        if not intent_candidates:
            result["status"] = FeedbackReprocessStatus.SKIPPED.value
            result["message"] = "Kh√¥ng t√¨m th·∫•y intent ph√π h·ª£p"
            results.append(result)
            continue

        selected_intent = None
        if gemini_service:
            try:
                selected_intent = gemini_service.select_best_intent(
                    feedback["feedback_text"],
                    intent_candidates
                )
            except Exception as exc:  # pragma: no cover - external dependency
                logger.error(
                    "Gemini selection error (feedback %s): %s",
                    feedback_id,
                    exc,
                    exc_info=True,
                )
        else:
            logger.warning(
                "Gemini service unavailable; skipping intent selection for feedback %s",
                feedback_id,
            )

        if not selected_intent:
            result["status"] = FeedbackReprocessStatus.SKIPPED.value
            result["message"] = "Gemini kh√¥ng ch·ªçn ƒë∆∞·ª£c intent"
            results.append(result)
            continue

        update_payload = FeedbackSentimentUpdate(
            level1_id=selected_intent["level1"]["id"],
            level2_id=selected_intent["level2"]["id"],
            level3_id=selected_intent["level3"]["id"],
            is_model_confirmed=False,
        )

        try:
            with get_db() as conn:
                updated_feedback = FeedbackSentimentCRUD.update(conn, feedback_id, update_payload)
        except ValueError as exc:
            result["status"] = FeedbackReprocessStatus.FAILED.value
            result["message"] = str(exc)
            results.append(result)
            continue
        except Exception as exc:
            logger.error(
                "Failed to update feedback %s after reprocess: %s",
                feedback_id,
                exc,
                exc_info=True,
            )
            result["status"] = FeedbackReprocessStatus.FAILED.value
            result["message"] = "Kh√¥ng c·∫≠p nh·∫≠t ƒë∆∞·ª£c feedback"
            results.append(result)
            continue

        if not updated_feedback:
            result["status"] = FeedbackReprocessStatus.FAILED.value
            result["message"] = "Feedback kh√¥ng t·ªìn t·∫°i khi c·∫≠p nh·∫≠t"
            results.append(result)
            continue

        try:
            manager = get_training_manager()
        except RuntimeError:
            manager = None
        except Exception as exc:  # pragma: no cover - defensive
            logger.error("Kh√¥ng l·∫•y ƒë∆∞·ª£c training manager: %s", exc)
            manager = None
        if manager:
            await manager.record_relabel_async("intent", updated_feedback)

        result["status"] = FeedbackReprocessStatus.UPDATED.value
        result["message"] = (
            f"Intent m·ªõi: {selected_intent['level1']['name']} ‚Üí "
            f"{selected_intent['level2']['name']} ‚Üí {selected_intent['level3']['name']}"
        )
        logger.info(
            "Reprocessed feedback %s with new intent (%s, %s, %s)",
            feedback_id,
            selected_intent["level1"]["id"],
            selected_intent["level2"]["id"],
            selected_intent["level3"]["id"],
        )
        results.append(result)

    return results


@router.post(
    "/feedbacks",
    response_model=FeedbackSentimentResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Submit feedback and analyze sentiment + intent"
)
async def create_feedback_sentiment(feedback_data: FeedbackSentimentCreate):
    """
    Submit customer feedback and get:
    1. Sentiment analysis (positive/negative/neutral)
    2. Intent classification using Gemini AI (level 1, 2, 3 labels)
    """
    try:
        # Step 1: Call sentiment service
        logger.info(f"Analyzing sentiment for feedback: {feedback_data.feedback_text[:50]}...")
        sentiment_result = await call_sentiment_service(feedback_data.feedback_text)
        
        sentiment_label = sentiment_result.get("label")
        confidence_score = sentiment_result.get("score")
        
        if not sentiment_label or confidence_score is None:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Invalid response from sentiment service"
            )
        
        # Step 2: Get embedding for feedback
        logger.info("Getting embedding for feedback...")
        feedback_embedding = await call_embedding_service(feedback_data.feedback_text)
        
        if not feedback_embedding:
            logger.warning("Failed to get embedding, saving feedback without intent")
            with get_db() as conn:
                feedback = FeedbackSentimentCRUD.create(
                    conn, feedback_data, sentiment_label, confidence_score,
                    is_model_confirmed=False
                )
                # Th·ª≠ t·ª± ƒë·ªông ph√¢n lo·∫°i l·∫°i sau khi l∆∞u
                try:
                    intent_result = await _classify_intent_for_feedback(feedback_data.feedback_text, feedback['id'])
                    if intent_result:
                        update_data = FeedbackSentimentUpdate(
                            level1_id=intent_result['level1_id'],
                            level2_id=intent_result['level2_id'],
                            level3_id=intent_result['level3_id'],
                            is_model_confirmed=False
                        )
                        feedback = FeedbackSentimentCRUD.update(conn, feedback['id'], update_data)
                        if feedback:
                            return feedback
                except Exception as e:
                    logger.warning(f"Failed to auto-classify intent after save: {e}")
                return feedback
        
        # Step 3: Get top 10 intent candidates (5 L1 ‚Üí 15 L2 ‚Üí 50 L3 ‚Üí 10 best)
        logger.info("Calculating top 10 intent candidates...")
        with get_db() as conn:
            intent_candidates = FeedbackIntentCRUD.get_top_intents(
                conn,
                feedback_embedding,
                limit=10,
                top_level1=5,
                top_level2_total=15,
                top_level3_total=50
            )
        
        if not intent_candidates:
            logger.warning("No intent candidates found, saving feedback without intent")
            with get_db() as conn:
                feedback = FeedbackSentimentCRUD.create(
                    conn, feedback_data, sentiment_label, confidence_score,
                    is_model_confirmed=False
                )
                return feedback
        
        # Step 4: Use Gemini to select best intent from top 10
        logger.info(f"Calling Gemini to select best intent from {len(intent_candidates)} candidates...")
        try:
            gemini_service = get_gemini_service()
            selected_intent = gemini_service.select_best_intent(
                feedback_data.feedback_text,
                intent_candidates
            )
        except Exception as e:
            logger.error(f"Gemini service error: {e}, continuing without intent")
            selected_intent = None
        
        # Step 5: Save to database with selected intent
        with get_db() as conn:
            if selected_intent:
                feedback = FeedbackSentimentCRUD.create(
                    conn,
                    feedback_data,
                    sentiment_label,
                    confidence_score,
                    level1_id=selected_intent['level1']['id'],
                    level2_id=selected_intent['level2']['id'],
                    level3_id=selected_intent['level3']['id'],
                    is_model_confirmed=False  # T·ª± ƒë·ªông ph√¢n lo·∫°i -> ch·ªù x√°c nh·∫≠n
                )
                logger.info(f"Feedback saved with intent: {selected_intent['level1']['name']} ‚Üí "
                          f"{selected_intent['level2']['name']} ‚Üí {selected_intent['level3']['name']}")
            else:
                feedback = FeedbackSentimentCRUD.create(
                    conn, feedback_data, sentiment_label, confidence_score,
                    is_model_confirmed=False
                )
                logger.warning("Feedback saved without intent (Gemini failed)")
            
            return feedback
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating feedback sentiment: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to create feedback"
        )


@router.put(
    "/feedbacks/{feedback_id}",
    response_model=FeedbackSentimentResponse,
    summary="Update sentiment or intent labels for a feedback"
)
def update_feedback_sentiment(feedback_id: UUID, update_data: FeedbackSentimentUpdate):
    """Manually update sentiment label and/or intent hierarchy for an existing feedback.
    
    Khi s·ª≠a b·∫±ng tay qua API PUT, t·ª± ƒë·ªông set is_model_confirmed = True (ƒë√£ x√°c nh·∫≠n).
    """
    try:
        with get_db() as conn:
            existing = FeedbackSentimentCRUD.get_by_id(conn, feedback_id)
            if not existing:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Feedback with id {feedback_id} not found"
                )
            
            # N·∫øu user s·ª≠a b·∫±ng tay v√† c√≥ thay ƒë·ªïi sentiment ho·∫∑c intent
            # nh∆∞ng kh√¥ng ch·ªâ ƒë·ªãnh is_model_confirmed, th√¨ m·∫∑c ƒë·ªãnh = True (ƒë√£ x√°c nh·∫≠n)
            update_dict = update_data.model_dump(exclude_unset=True)
            has_sentiment_change = "sentiment_label" in update_dict
            has_intent_change = any(key in update_dict for key in ["level1_id", "level2_id", "level3_id"])
            
            if (has_sentiment_change or has_intent_change) and "is_model_confirmed" not in update_dict:
                # S·ª≠a b·∫±ng tay qua API -> t·ª± ƒë·ªông x√°c nh·∫≠n
                update_data.is_model_confirmed = True
            
            updated_feedback = FeedbackSentimentCRUD.update(conn, feedback_id, update_data)
            if not updated_feedback:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Feedback with id {feedback_id} not found"
                )
        return updated_feedback
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating feedback {feedback_id}: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to update feedback"
        )


@router.post(
    "/feedbacks/{feedback_id}/confirm",
    response_model=FeedbackSentimentResponse,
    summary="Confirm model prediction for a feedback"
)
def confirm_feedback_sentiment(feedback_id: UUID):
    """ƒê√°nh d·∫•u feedback ƒë√£ ƒë∆∞·ª£c ng∆∞·ªùi d√πng x√°c nh·∫≠n m√¥ h√¨nh d·ª± ƒëo√°n ƒë√∫ng."""
    try:
        with get_db() as conn:
            existing = FeedbackSentimentCRUD.get_by_id(conn, feedback_id)
            payload = FeedbackSentimentUpdate(is_model_confirmed=True)
            updated_feedback = FeedbackSentimentCRUD.update(conn, feedback_id, payload)
            if not updated_feedback:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Feedback with id {feedback_id} not found"
                )
        return updated_feedback
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error confirming feedback {feedback_id}: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to confirm feedback"
        )


@router.get(
    "/feedbacks",
    response_model=FeedbackSentimentListResponse,
    summary="Get all feedback sentiments with filters"
)
def get_feedback_sentiments(
    sentiment_label: Optional[SentimentLabel] = Query(None, description="Filter by sentiment label"),
    feedback_source: Optional[FeedbackSource] = Query(None, description="Filter by feedback source"),
    skip: int = Query(0, ge=0, description="Number of records to skip"),
    limit: int = Query(100, ge=1, le=1000, description="Maximum number of records"),
):
    """Get all feedback sentiments with optional filters."""
    try:
        with get_db() as conn:
            # Convert enums to values for database query
            label_value = sentiment_label.value if sentiment_label else None
            source_value = feedback_source.value if feedback_source else None
            
            feedbacks = FeedbackSentimentCRUD.get_all(
                conn,
                sentiment_label=label_value,
                feedback_source=source_value,
                skip=skip,
                limit=limit
            )
            total = FeedbackSentimentCRUD.count(
                conn,
                sentiment_label=label_value,
                feedback_source=source_value
            )
            
            return FeedbackSentimentListResponse(
                feedbacks=feedbacks,
                total=total,
                sentiment_label=sentiment_label,
                feedback_source=feedback_source
            )
    except Exception as e:
        logger.error(f"Error getting feedback sentiments: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to retrieve feedbacks"
        )


@router.get(
    "/feedbacks/{feedback_id}",
    response_model=FeedbackSentimentResponse,
    summary="Get feedback sentiment by ID"
)
def get_feedback_sentiment(feedback_id: UUID):
    """Get a specific feedback sentiment by ID."""
    try:
        with get_db() as conn:
            feedback = FeedbackSentimentCRUD.get_by_id(conn, feedback_id)
            if not feedback:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Feedback with id {feedback_id} not found"
                )
            return feedback
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting feedback {feedback_id}: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to retrieve feedback"
        )


@router.post(
    "/feedbacks/import",
    response_model=FeedbackImportResponse,
    summary="Import feedback data from Excel file"
)
async def import_feedbacks_from_excel(file: UploadFile = File(...)):
    """Import feedback sentiments from an Excel file."""
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Ch·ªâ h·ªó tr·ª£ import file Excel ƒë·ªãnh d·∫°ng .xlsx"
        )

    content = await file.read()
    if not content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File Excel kh√¥ng c√≥ d·ªØ li·ªáu"
        )

    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as exc:  # pragma: no cover - library-specific error
        logger.error(f"Failed to read Excel file during import: {exc}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Kh√¥ng th·ªÉ ƒë·ªçc file Excel. Vui l√≤ng ki·ªÉm tra ƒë·ªãnh d·∫°ng v√† n·ªôi dung."
        )

    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File Excel kh√¥ng c√≥ d√≤ng header"
        )

    normalized_headers = [
        str(cell).strip().lower() if cell is not None else ""
        for cell in header_row
    ]
    required_headers = ["content", "sentiment", "level1", "level2", "level3"]
    missing_columns = [col for col in required_headers if col not in normalized_headers]
    if missing_columns:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"File Excel thi·∫øu c√°c c·ªôt b·∫Øt bu·ªôc: {', '.join(missing_columns)}"
        )

    header_index = {name: idx for idx, name in enumerate(normalized_headers)}

    def _normalize(value) -> str:
        if value is None:
            return ""
        return str(value).strip()

    inserted = 0
    error_rows = []

    with get_db() as conn:
        labels = LabelCRUD.get_all(conn, skip=0, limit=100000)
        labels_by_level = {1: {}, 2: {}, 3: {}}
        for label in labels:
            label_name = _normalize(label.get("name"))
            if label_name:
                labels_by_level[label["level"]][label_name.lower()] = label

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_values = list(row) if row else []
            if len(row_values) < len(normalized_headers):
                row_values.extend([None] * (len(normalized_headers) - len(row_values)))

            if all(
                _normalize(row_values[header_index[col]]) == ""
                for col in required_headers
            ):
                continue  # Skip empty row

            def _get_value(column: str):
                idx = header_index[column]
                if idx >= len(row_values):
                    return None
                return row_values[idx]

            raw_content = _get_value("content")
            raw_sentiment = _get_value("sentiment")
            raw_level1 = _get_value("level1")
            raw_level2 = _get_value("level2")
            raw_level3 = _get_value("level3")

            content_value = _normalize(raw_content)
            sentiment_value = _normalize(raw_sentiment)
            level1_value = _normalize(raw_level1)
            level2_value = _normalize(raw_level2)
            level3_value = _normalize(raw_level3)

            row_errors = []
            if not content_value:
                row_errors.append("Thi·∫øu n·ªôi dung feedback")

            sentiment_enum = None
            if not sentiment_value:
                row_errors.append("Thi·∫øu gi√° tr·ªã sentiment")
            else:
                try:
                    sentiment_enum = SentimentLabel(sentiment_value.upper())
                except ValueError:
                    row_errors.append(f"Sentiment kh√¥ng h·ª£p l·ªá: {sentiment_value}")

            level1_label = None
            level2_label = None
            level3_label = None

            if level1_value:
                level1_label = labels_by_level[1].get(level1_value.lower())
                if not level1_label:
                    row_errors.append(f"Level1 '{level1_value}' kh√¥ng t·ªìn t·∫°i trong h·ªá th·ªëng")
            elif level2_value or level3_value:
                row_errors.append("Ph·∫£i cung c·∫•p Level1 khi c√≥ Level2/Level3")

            if level2_value:
                if not level1_label:
                    row_errors.append("Level2 y√™u c·∫ßu Level1 h·ª£p l·ªá")
                else:
                    level2_label = labels_by_level[2].get(level2_value.lower())
                    if not level2_label:
                        row_errors.append(f"Level2 '{level2_value}' kh√¥ng t·ªìn t·∫°i trong h·ªá th·ªëng")
                    elif level2_label["parent_id"] != level1_label["id"]:
                        row_errors.append(
                            f"Level2 '{level2_value}' kh√¥ng thu·ªôc Level1 '{level1_label['name']}'"
                        )

            if level3_value:
                if not level2_label:
                    row_errors.append("Level3 y√™u c·∫ßu Level2 h·ª£p l·ªá")
                else:
                    level3_label = labels_by_level[3].get(level3_value.lower())
                    if not level3_label:
                        row_errors.append(f"Level3 '{level3_value}' kh√¥ng t·ªìn t·∫°i trong h·ªá th·ªëng")
                    elif level3_label["parent_id"] != level2_label["id"]:
                        row_errors.append(
                            f"Level3 '{level3_value}' kh√¥ng thu·ªôc Level2 '{level2_label['name']}'"
                        )

            # N·∫øu c√≥ l·ªói v·ªÅ sentiment, b·ªè qua d√≤ng n√†y
            if sentiment_enum is None:
                error_rows.append(
                    {
                        "row_index": row_idx,
                        "content": content_value,
                        "sentiment": sentiment_value,
                        "level1": level1_value,
                        "level2": level2_value,
                        "level3": level3_value,
                        "errors": row_errors or ["Kh√¥ng x√°c ƒë·ªãnh ƒë∆∞·ª£c sentiment"],
                    }
                )
                continue

            feedback_data = FeedbackSentimentCreate(
                feedback_text=content_value,
                feedback_source=FeedbackSource.WEB,
            )

            level1_id = level1_label["id"] if level1_label else None
            level2_id = level2_label["id"] if level2_label else None
            level3_id = level3_label["id"] if level3_label else None
            
            # Ki·ªÉm tra xem c√≥ ƒë·ªß nh√£n v√† t·∫•t c·∫£ ƒë·ªÅu h·ª£p l·ªá kh√¥ng
            # C·∫ßn c√≥ ƒë·ªß c·∫£ 3 nh√£n (level1, level2, level3) v√† kh√¥ng c√≥ l·ªói n√†o
            has_all_labels = level1_id is not None and level2_id is not None and level3_id is not None
            has_no_label_errors = not any("Level" in error or "level" in error.lower() for error in row_errors)
            
            if has_all_labels and has_no_label_errors:
                # C√≥ ƒë·ªß nh√£n v√† t·∫•t c·∫£ ƒë·ªÅu h·ª£p l·ªá -> ƒë√£ ƒë∆∞·ª£c x√°c nh·∫≠n (is_model_confirmed = True)
                is_confirmed = True
                logger.info(f"Row {row_idx}: C√≥ ƒë·ªß nh√£n h·ª£p l·ªá, l∆∞u v·ªõi is_model_confirmed = True")
            else:
                # Thi·∫øu nh√£n ho·∫∑c c√≥ l·ªói v·ªÅ nh√£n -> t·ª± ƒë·ªông ph√¢n lo·∫°i l·∫°i
                logger.info(f"Row {row_idx}: Thi·∫øu nh√£n ho·∫∑c nh√£n kh√¥ng h·ª£p l·ªá, s·∫Ω t·ª± ƒë·ªông ph√¢n lo·∫°i intent")
                intent_result = await _classify_intent_for_feedback(content_value)
                if intent_result:
                    level1_id = intent_result['level1_id']
                    level2_id = intent_result['level2_id']
                    level3_id = intent_result['level3_id']
                    # T·ª± ƒë·ªông ph√¢n lo·∫°i -> ch·ªù x√°c nh·∫≠n (is_model_confirmed = False)
                    is_confirmed = False
                else:
                    # Kh√¥ng th·ªÉ ph√¢n lo·∫°i -> l∆∞u kh√¥ng c√≥ intent
                    level1_id = None
                    level2_id = None
                    level3_id = None
                    is_confirmed = False

            FeedbackSentimentCRUD.create(
                conn,
                feedback_data,
                sentiment_enum.value,
                confidence_score=1.0,
                level1_id=level1_id,
                level2_id=level2_id,
                level3_id=level3_id,
                is_model_confirmed=is_confirmed,
            )
            inserted += 1

    log_file_path = None
    if error_rows:
        log_dir = Path(__file__).resolve().parent / "logs" / "feedback_import"
        log_dir.mkdir(parents=True, exist_ok=True)
        log_file_path = log_dir / f"errors_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
        with log_file_path.open("w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(["row_index", "content", "sentiment", "level1", "level2", "level3", "errors"])
            for row in error_rows:
                writer.writerow(
                    [
                        row["row_index"],
                        row["content"],
                        row["sentiment"],
                        row["level1"],
                        row["level2"],
                        row["level3"],
                        " | ".join(row["errors"]),
                    ]
                )

    log_file_str = None
    if log_file_path:
        try:
            log_file_str = str(log_file_path.relative_to(Path(__file__).resolve().parent))
        except ValueError:
            log_file_str = str(log_file_path)

    return FeedbackImportResponse(
        imported=inserted,
        failed=len(error_rows),
        log_file=log_file_str,
    )


@router.post(
    "/feedbacks/import-simple",
    response_model=FeedbackImportResponse,
    summary="Import feedback data from Excel file (simple format: content and source only)"
)
async def import_feedbacks_simple_from_excel(file: UploadFile = File(...)):
    """
    Import feedback sentiments from an Excel file with simple format.
    Required columns: content (n·ªôi dung feedback)
    Optional columns: source (ngu·ªìn feedback, default: 'web')
    The system will automatically analyze sentiment and intent for each feedback.
    """
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Ch·ªâ h·ªó tr·ª£ import file Excel ƒë·ªãnh d·∫°ng .xlsx"
        )

    content = await file.read()
    if not content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File Excel kh√¥ng c√≥ d·ªØ li·ªáu"
        )

    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as exc:
        logger.error(f"Failed to read Excel file during import: {exc}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Kh√¥ng th·ªÉ ƒë·ªçc file Excel. Vui l√≤ng ki·ªÉm tra ƒë·ªãnh d·∫°ng v√† n·ªôi dung."
        )

    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File Excel kh√¥ng c√≥ d√≤ng header"
        )

    normalized_headers = [
        str(cell).strip().lower() if cell is not None else ""
        for cell in header_row
    ]
    
    # T√¨m c√°c c·ªôt c·∫ßn thi·∫øt
    content_col_idx = None
    source_col_idx = None
    
    for idx, header in enumerate(normalized_headers):
        if header in ["content", "n·ªôi dung", "feedback", "feedback_text", "text"]:
            content_col_idx = idx
        elif header in ["source", "ngu·ªìn", "feedback_source"]:
            source_col_idx = idx
    
    if content_col_idx is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File Excel ph·∫£i c√≥ c·ªôt 'content' (ho·∫∑c 'n·ªôi dung', 'feedback', 'feedback_text', 'text')"
        )

    inserted = 0
    error_rows = []

    def _normalize(value) -> str:
        if value is None:
            return ""
        return str(value).strip()

    # X·ª≠ l√Ω t·ª´ng d√≤ng
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row_values = list(row) if row else []
        
        # L·∫•y n·ªôi dung feedback
        if content_col_idx >= len(row_values):
            continue
        content_value = _normalize(row_values[content_col_idx])
        
        if not content_value:
            continue  # B·ªè qua d√≤ng tr·ªëng
        
        # L·∫•y ngu·ªìn feedback (m·∫∑c ƒë·ªãnh l√† 'web')
        source_value = 'web'
        if source_col_idx is not None and source_col_idx < len(row_values):
            raw_source = _normalize(row_values[source_col_idx])
            if raw_source:
                # Chuy·ªÉn ƒë·ªïi t√™n ngu·ªìn sang gi√° tr·ªã enum
                source_map = {
                    'web': 'web',
                    'app': 'app',
                    'map': 'map',
                    'form kh·∫£o s√°t': 'form kh·∫£o s√°t',
                    't·ªïng ƒë√†i': 't·ªïng ƒë√†i',
                }
                source_value = source_map.get(raw_source.lower(), 'web')
        
        # T·∫°o feedback v√† t·ª± ƒë·ªông ph√¢n t√≠ch
        try:
            feedback_data = FeedbackSentimentCreate(
                feedback_text=content_value,
                feedback_source=FeedbackSource(source_value)
            )
            
            # S·ª≠ d·ª•ng logic t∆∞∆°ng t·ª± nh∆∞ create_feedback_sentiment
            # Step 1: Call sentiment service
            sentiment_result = await call_sentiment_service(feedback_data.feedback_text)
            sentiment_label = sentiment_result.get("label")
            confidence_score = sentiment_result.get("score")
            
            if not sentiment_label or confidence_score is None:
                error_rows.append({
                    "row_index": row_idx,
                    "content": content_value,
                    "errors": ["Kh√¥ng th·ªÉ ph√¢n t√≠ch sentiment"]
                })
                continue
            
            # Step 2: Get embedding for feedback
            feedback_embedding = await call_embedding_service(feedback_data.feedback_text)
            
            level1_id = None
            level2_id = None
            level3_id = None
            
            if feedback_embedding:
                # Step 3: Get top 10 intent candidates
                with get_db() as conn:
                    intent_candidates = FeedbackIntentCRUD.get_top_intents(
                        conn,
                        feedback_embedding,
                        limit=10,
                        top_level1=5,
                        top_level2_total=15,
                        top_level3_total=50
                    )
                    
                    if intent_candidates:
                        # Step 4: Use Gemini to select best intent
                        try:
                            gemini_service = get_gemini_service()
                            selected_intent = gemini_service.select_best_intent(
                                feedback_data.feedback_text,
                                intent_candidates
                            )
                            if selected_intent:
                                level1_id = selected_intent['level1']['id']
                                level2_id = selected_intent['level2']['id']
                                level3_id = selected_intent['level3']['id']
                        except Exception as e:
                            logger.warning(f"Gemini service error for row {row_idx}: {e}, continuing without intent")
                
                # Step 5: Save to database
                # T·ª± ƒë·ªông ph√¢n lo·∫°i -> ch·ªù x√°c nh·∫≠n (is_model_confirmed = False)
                with get_db() as conn:
                    FeedbackSentimentCRUD.create(
                        conn,
                        feedback_data,
                        sentiment_label,
                        confidence_score,
                        level1_id=level1_id,
                        level2_id=level2_id,
                        level3_id=level3_id,
                        is_model_confirmed=False  # T·ª± ƒë·ªông ph√¢n lo·∫°i -> ch·ªù x√°c nh·∫≠n
                    )
            else:
                # Save without intent if embedding fails
                with get_db() as conn:
                    FeedbackSentimentCRUD.create(
                        conn,
                        feedback_data,
                        sentiment_label,
                        confidence_score,
                        is_model_confirmed=False
                    )
            
            inserted += 1
            
        except Exception as e:
            logger.error(f"Error processing row {row_idx}: {e}", exc_info=True)
            error_rows.append({
                "row_index": row_idx,
                "content": content_value,
                "errors": [str(e)]
            })

    # Ghi log file n·∫øu c√≥ l·ªói
    log_file_path = None
    if error_rows:
        log_dir = Path(__file__).resolve().parent / "logs" / "feedback_import"
        log_dir.mkdir(parents=True, exist_ok=True)
        log_file_path = log_dir / f"errors_simple_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
        with log_file_path.open("w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(["row_index", "content", "errors"])
            for row in error_rows:
                writer.writerow([
                    row["row_index"],
                    row["content"],
                    " | ".join(row["errors"]),
                ])

    log_file_str = None
    if log_file_path:
        try:
            log_file_str = str(log_file_path.relative_to(Path(__file__).resolve().parent))
        except ValueError:
            log_file_str = str(log_file_path)

    return FeedbackImportResponse(
        imported=inserted,
        failed=len(error_rows),
        log_file=log_file_str,
    )


# --- Intent Analysis Routes ---

@router.post(
    "/feedbacks/{feedback_id}/intents",
    response_model=FeedbackIntentResponse,
    status_code=status.HTTP_200_OK,
    summary="Analyze feedback intents"
)
async def analyze_feedback_intents(feedback_id: UUID):
    """
    Analyze feedback and return top 10 intent triplets based on cosine similarity.
    This endpoint calculates or retrieves cached intent analysis results.
    """
    try:
        with get_db() as conn:
            # Check if feedback exists
            feedback = FeedbackSentimentCRUD.get_by_id(conn, feedback_id)
            if not feedback:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Feedback with id {feedback_id} not found"
                )
            
            # Try to get cached intents first (top 50)
            cached_intents = FeedbackIntentCRUD.get_cached_intents(conn, feedback_id, limit=50)
            
            if cached_intents:
                logger.info(f"Returning cached intents for feedback {feedback_id}")
                return FeedbackIntentResponse(
                    feedback_id=feedback_id,
                    intents=[IntentTriplet(**intent) for intent in cached_intents],
                    total_intents=len(cached_intents)
                )
            
            # Calculate new intents
            logger.info(f"Calculating new intents for feedback {feedback_id}")
            
            # Get embedding for feedback text
            feedback_embedding = await call_embedding_service(feedback['feedback_text'])
            
            if not feedback_embedding:
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail="Failed to get feedback embedding"
                )
            
            # Calculate top intents using hierarchical approach
            # Top 5 level1 ‚Üí top 20 level2 ‚Üí top 50 triplets
            intents = FeedbackIntentCRUD.get_top_intents(
                conn, 
                feedback_embedding, 
                limit=50,
                top_level1=5,
                top_level2_per_level1=4,
                top_level3_per_level2=3
            )
            
            if not intents:
                logger.warning(f"No intents found for feedback {feedback_id}")
                return FeedbackIntentResponse(
                    feedback_id=feedback_id,
                    intents=[],
                    total_intents=0
                )
            
            # Save intents to cache
            FeedbackIntentCRUD.save_intents(conn, feedback_id, intents)
            
            return FeedbackIntentResponse(
                feedback_id=feedback_id,
                intents=[IntentTriplet(**intent) for intent in intents],
                total_intents=len(intents)
            )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error analyzing intents for feedback {feedback_id}: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to analyze feedback intents"
        )


@router.post(
    "/feedbacks/auto-classify-intents",
    summary="T·ª± ƒë·ªông ph√¢n lo·∫°i intent cho c√°c feedback ch∆∞a c√≥ intent"
)
async def auto_classify_feedbacks_without_intents():
    """
    T·ª± ƒë·ªông ph√¢n lo·∫°i intent cho t·∫•t c·∫£ feedback ch∆∞a c√≥ intent (level1_id IS NULL).
    Qu√©t to√†n b·ªô database v√† ph√¢n lo·∫°i intent cho c√°c feedback ch∆∞a ƒë∆∞·ª£c ph√¢n lo·∫°i.
    C√°c feedback sau khi ƒë∆∞·ª£c ph√¢n lo·∫°i s·∫Ω c√≥ is_model_confirmed = False (ch·ªù x√°c nh·∫≠n).
    """
    try:
        with get_db() as conn:
            from database import execute_query
            
            # L·∫•y t·∫•t c·∫£ feedback ch∆∞a c√≥ intent (level1_id, level2_id, level3_id ƒë·ªÅu NULL)
            results = execute_query(
                conn,
                """
                SELECT id, feedback_text 
                FROM feedback_sentiments 
                WHERE level1_id IS NULL 
                  AND level2_id IS NULL 
                  AND level3_id IS NULL
                ORDER BY created_at DESC
                """,
                fetch="all"
            )
            
            feedbacks = [dict(row) for row in results] if results else []
            
            if not feedbacks:
                return {
                    "status": "success",
                    "message": "Kh√¥ng c√≥ feedback n√†o c·∫ßn ph√¢n lo·∫°i",
                    "total": 0,
                    "classified": 0,
                    "failed": 0
                }
            
            # Ki·ªÉm tra xem labels ƒë√£ c√≥ embeddings ch∆∞a
            labels_with_embeddings = LabelCRUD.get_all_with_embeddings(conn)
            has_all_levels = (
                len(labels_with_embeddings.get(1, [])) > 0 and
                len(labels_with_embeddings.get(2, [])) > 0 and
                len(labels_with_embeddings.get(3, [])) > 0
            )
            
            if not has_all_levels:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Labels ch∆∞a c√≥ embeddings. Vui l√≤ng ch·∫°y endpoint POST /admin/seed-label-embeddings tr∆∞·ªõc ƒë·ªÉ seed embeddings cho labels."
                )
            
            logger.info(f"B·∫Øt ƒë·∫ßu t·ª± ƒë·ªông ph√¢n lo·∫°i intent cho {len(feedbacks)} feedback")
            
            classified = 0
            failed = 0
            
            for feedback in feedbacks:
                try:
                    feedback_id = feedback['id']
                    feedback_text = feedback['feedback_text']
                    
                    # Ph√¢n lo·∫°i intent
                    intent_result = await _classify_intent_for_feedback(feedback_text, feedback_id)
                    
                    if intent_result:
                        # C·∫≠p nh·∫≠t feedback v·ªõi intent m·ªõi, is_model_confirmed = False
                        update_data = FeedbackSentimentUpdate(
                            level1_id=intent_result['level1_id'],
                            level2_id=intent_result['level2_id'],
                            level3_id=intent_result['level3_id'],
                            is_model_confirmed=False
                        )
                        FeedbackSentimentCRUD.update(conn, feedback_id, update_data)
                        classified += 1
                        logger.info(f"ƒê√£ ph√¢n lo·∫°i intent cho feedback {feedback_id}")
                    else:
                        failed += 1
                        logger.warning(f"Kh√¥ng th·ªÉ ph√¢n lo·∫°i intent cho feedback {feedback_id}")
                        
                except Exception as e:
                    failed += 1
                    logger.error(f"L·ªói khi ph√¢n lo·∫°i feedback {feedback.get('id')}: {e}", exc_info=True)
            
            return {
                "status": "success",
                "message": f"ƒê√£ ph√¢n lo·∫°i {classified} feedback, {failed} feedback th·∫•t b·∫°i",
                "total": len(feedbacks),
                "classified": classified,
                "failed": failed
            }
    
    except Exception as e:
        logger.error(f"Error auto-classifying feedbacks: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Kh√¥ng th·ªÉ t·ª± ƒë·ªông ph√¢n lo·∫°i feedback: {str(e)}"
        )


@router.get(
    "/feedbacks/{feedback_id}/intents",
    response_model=FeedbackIntentResponse,
    summary="Get cached feedback intents"
)
def get_feedback_intents(feedback_id: UUID):
    """Get cached intent analysis results for a feedback."""
    try:
        with get_db() as conn:
            # Check if feedback exists
            feedback = FeedbackSentimentCRUD.get_by_id(conn, feedback_id)
            if not feedback:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Feedback with id {feedback_id} not found"
                )
            
            # Get cached intents (top 50)
            cached_intents = FeedbackIntentCRUD.get_cached_intents(conn, feedback_id, limit=50)
            
            if not cached_intents:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"No intent analysis found for feedback {feedback_id}. Please use POST endpoint to analyze."
                )
            
            return FeedbackIntentResponse(
                feedback_id=feedback_id,
                intents=[IntentTriplet(**intent) for intent in cached_intents],
                total_intents=len(cached_intents)
            )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting intents for feedback {feedback_id}: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to retrieve feedback intents"
        )


# --- Admin/Seed Data Routes ---

@router.post(
    "/admin/seed-label-embeddings",
    summary="Seed embeddings for all labels (Admin)"
)
async def seed_label_embeddings():
    """
    Admin endpoint to compute and update embeddings for all labels.
    This should be called after adding new labels or when initializing the system.
    """
    try:
        with get_db() as conn:
            # Get all labels
            all_labels = LabelCRUD.get_all(conn, skip=0, limit=1000)
            
            if not all_labels:
                return {
                    "status": "success",
                    "message": "No labels found to process",
                    "total": 0,
                    "processed": 0,
                    "failed": 0
                }
            
            logger.info(f"Starting to seed embeddings for {len(all_labels)} labels")
            
            processed = 0
            failed = 0
            
            for label in all_labels:
                try:
                    # Create text for embedding
                    text = label['name']
                    if label.get('description'):
                        text = f"{label['name']}. {label['description']}"
                    
                    # Get embedding
                    embedding = await call_embedding_service(text)
                    
                    if not embedding:
                        logger.warning(f"Failed to get embedding for label: {label['name']} (id={label['id']})")
                        failed += 1
                        continue
                    
                    # Update label with embedding
                    LabelCRUD.update_embedding(conn, label['id'], embedding)
                    processed += 1
                    logger.debug(f"Updated embedding for label: {label['name']} (id={label['id']})")
                    
                except Exception as e:
                    logger.error(f"Error processing label {label['id']}: {e}")
                    failed += 1
            
            return {
                "status": "success",
                "message": f"Completed seeding embeddings",
                "total": len(all_labels),
                "processed": processed,
                "failed": failed
            }
    
    except Exception as e:
        logger.error(f"Error in seed_label_embeddings: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to seed label embeddings: {str(e)}"
        )


@router.post(
    "/admin/seed-feedback-intents",
    summary="Seed intents for all feedbacks (Admin)"
)
async def seed_feedback_intents(recompute: bool = False):
    """
    Admin endpoint to compute and cache intents for all feedbacks.
    
    Args:
        recompute: If True, recompute intents for all feedbacks including those with cached results.
                  If False, only compute for feedbacks without cached results.
    """
    try:
        with get_db() as conn:
            # Get feedbacks to process
            if recompute:
                # Get all feedbacks
                from database import execute_query
                results = execute_query(
                    conn,
                    "SELECT id, feedback_text, created_at FROM feedback_sentiments ORDER BY created_at DESC",
                    fetch="all"
                )
                feedbacks = [dict(row) for row in results] if results else []
            else:
                # Get only feedbacks without intents
                from database import execute_query
                results = execute_query(
                    conn,
                    """
                    SELECT DISTINCT fs.id, fs.feedback_text, fs.created_at
                    FROM feedback_sentiments fs
                    LEFT JOIN feedback_intents fi ON fs.id = fi.feedback_id
                    WHERE fi.id IS NULL
                    ORDER BY fs.created_at DESC
                    """,
                    fetch="all"
                )
                feedbacks = [dict(row) for row in results] if results else []
            
            if not feedbacks:
                return {
                    "status": "success",
                    "message": "No feedbacks found to process",
                    "total": 0,
                    "processed": 0,
                    "failed": 0
                }
            
            logger.info(f"Starting to seed intents for {len(feedbacks)} feedbacks")
            
            processed = 0
            failed = 0
            
            for feedback in feedbacks:
                try:
                    feedback_id = feedback['id']
                    feedback_text = feedback['feedback_text']
                    
                    # Get embedding for feedback
                    feedback_embedding = await call_embedding_service(feedback_text)
                    
                    if not feedback_embedding:
                        logger.warning(f"Failed to get embedding for feedback {feedback_id}")
                        failed += 1
                        continue
                    
                    # Calculate top intents using hierarchical approach
                    # Top 5 level1 ‚Üí top 20 level2 ‚Üí top 50 triplets
                    intents = FeedbackIntentCRUD.get_top_intents(
                        conn, 
                        feedback_embedding, 
                        limit=50,
                        top_level1=5,
                        top_level2_per_level1=4,
                        top_level3_per_level2=3
                    )
                    
                    if intents:
                        # Save intents to cache
                        FeedbackIntentCRUD.save_intents(conn, feedback_id, intents)
                        processed += 1
                        logger.debug(f"Computed intents for feedback {feedback_id}")
                    else:
                        logger.warning(f"No intents found for feedback {feedback_id}")
                        failed += 1
                    
                except Exception as e:
                    logger.error(f"Error processing feedback {feedback.get('id')}: {e}")
                    failed += 1
            
            return {
                "status": "success",
                "message": f"Completed seeding intents",
                "total": len(feedbacks),
                "processed": processed,
                "failed": failed,
                "recompute": recompute
            }
    
    except Exception as e:
        logger.error(f"Error in seed_feedback_intents: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to seed feedback intents: {str(e)}"
        )
